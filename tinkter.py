#!/usr/bin/env python
# coding: utf-8

#!/usr/bin/env python
# coding: utf-8

#Import Library

import tkinter as tk
from tkinter import ttk  
from tkinter import filedialog
from tkinter import *
from tkinter.ttk import * 
import tkinter.messagebox as mbox

from ibapi.client import EClient
from ibapi.wrapper import EWrapper
from ibapi.contract import Contract
from ibapi.order import *
import math
import numpy as np
import pandas as pd
import os
import openpyxl as excel

import time
from dateutil.parser import parse
from datetime import datetime
from datetime import date
from dateutil.relativedelta import relativedelta
from threading import Timer

import warnings
warnings.filterwarnings('ignore')


#Utility Functions

counter=0
def execute_method():
    print("OK")
    

def center_window(w=300, h=200):
    # get screen width and height
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    
def init_output_file(path,flag):
    if flag==1:
        wb = excel.Workbook()
        wb.save(filename=path)
        for sheetName in wb.sheetnames:
            del wb[sheetName]
            
    if flag==2:        
        wb = excel.load_workbook(path)
        std=wb['Sheet']
        wb.remove(std)
        wb.save(filename=path)
        
def output_writer_headers(path,res,name,remove):
        book = excel.load_workbook(path)
        if remove==1:
            std=book[name]
            book.remove(std)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book
        res.to_excel(writer, sheet_name=name, index=False, header=False)
        writer.save()   

def output_writer(path,res,name,remove):
        book = excel.load_workbook(path)
        if remove==1:
            std=book[name]
            book.remove(std)
        writer = pd.ExcelWriter(path, engine = 'openpyxl')
        writer.book = book
        res.to_excel(writer, sheet_name=name, index=False)
        writer.save()


def check_int_float(s):
    res = 0
    try:
        res = int(s)
        return res
    except:
        res = 0
    try:
        res = float(s)
        return res
    except:
        res = 0
    return "False"


def new_box(map):
    r = Tk()
    height=len(map) * 100
    width = 800
    frame = Frame(r, width=width, height=height)
    frame.pack()
    width = 0
    for key, value in map.items():
        txt = key + " = " + str(value)
        if len(txt) > width:
            width = len(txt)
    for key,value in map.items():
        txt=key+ " = " +str(value)
        lab = Label(frame,text=txt,anchor=W, justify=LEFT,width=width+5)
        lab.pack()
    MyButton1 = Button(r, text="Close", width=10,style = 'W.TButton',command=r.destroy)
    MyButton1.pack()
    root.mainloop()
    
def hist_data_formating():
    var=20
    complete=pd.ExcelFile('Complete_Output.xlsx')
    M3T1DI=pd.ExcelFile('M3T1DI.xlsx')
    map={}
    for sheet in M3T1DI.sheet_names:
        counter=0
        data_file_complete = pd.read_excel(M3T1DI,sheet)
        print(sheet)
        for i in range(1,data_file_complete.shape[0]):
            range_p=round(data_file_complete.Close.iloc[i-1] * ((100+var)/100),2)
            range_n=round(data_file_complete.Close.iloc[i-1] * ((100-var)/100),2)
            if data_file_complete.Close.iloc[i] > range_p or data_file_complete.Close.iloc[i] < range_n:
                if i==1 and i!=data_file_complete.shape[0]:
                    range_p=round(data_file_complete.Close.iloc[i] * ((100+var)/100),2)
                    range_n=round(data_file_complete.Close.iloc[i] * ((100-var)/100),2)
                    if data_file_complete.Close.iloc[i+1] > range_p or data_file_complete.Close.iloc[i+1] < range_n:
                        temp=round((data_file_complete.Close.iloc[i-1]+data_file_complete.Close.iloc[i+1])/2,2)
                        if(temp!=data_file_complete.Close.iloc[i]):
                            counter=counter+1
                            print(counter,i,data_file_complete.Close.iloc[i],temp)
                            data_file_complete.Close.iloc[i]=temp
                elif  i==data_file_complete.shape[0]-1 :
                    range_p=round(data_file_complete.Close.iloc[i-1] * ((100+var)/100),2)
                    range_n=round(data_file_complete.Close.iloc[i-1] * ((100-var)/100),2)
                    if data_file_complete.Close.iloc[i] > range_p:
                        temp=range_p
                        if(temp!=data_file_complete.Close.iloc[i]):
                            counter=counter+1
                            print(counter,i,data_file_complete.Close.iloc[i],temp)
                            data_file_complete.Close.iloc[i]=temp
                    elif  data_file_complete.Close.iloc[i] < range_n:
                        temp=range_n
                        if(temp!=data_file_complete.Close.iloc[i]):
                            counter=counter+1
                            print(counter,i,data_file_complete.Close.iloc[i],temp)
                            data_file_complete.Close.iloc[i]=temp
                else:
                    temp=round((data_file_complete.Close.iloc[i-1]+data_file_complete.Close.iloc[i+1])/2,2)
                    if(temp!=data_file_complete.Close.iloc[i]):
                        counter=counter+1
                        print(counter,i,data_file_complete.Close.iloc[i],temp)
                        data_file_complete.Close.iloc[i]=temp
        map[sheet]=counter
        output_writer('M3T1DI.xlsx',data_file_complete,sheet,1)
    print("Data Check Successful !!")
    new_box(map)
    

    
def create_pi_file(filename):
    pi = pd.read_excel("M3T1PI.xlsx",sheet_name=None, encoding = 'unicode_escape', header=None)
    port_on=pi['Portfolio Ongoing']
    #net liq val
    portfolio1=pd.read_excel("m3t2OD2.xlsx", header=0)
    port_on.iloc[2:3,4:5]=portfolio1.loc[portfolio1["Key"] == "NetLiquidation","Value"].values[0]
    #open positions
    portfolio2=pd.read_excel("m3t2OD1.xlsx", header=0)
    symbol=portfolio2.Symbol
    position=portfolio2.Position

    for i in range(0,len(symbol)):
        
        port_on=port_on.append(pd.Series(dtype="float"), ignore_index=True)
        port_on.iloc[13+i:14+i,1:2]=symbol.values[i]
        port_on.iloc[13+i:14+i,2:3]=position.values[i]

    output_writer_headers(filename,port_on,'Portfolio Ongoing',1)
    print("PI File Created")


def clean_PI_file():
    pi = pd.read_excel("M3T1PI.xlsx", sheet_name=None, encoding='unicode_escape', header=None)
    path = 'Historical_data_inputs.xlsx'
    df = pd.read_excel(path, sheet_name="Sheet1", encoding='unicode_escape')
    input_symbol_hist = []
    for index, row in df.iterrows():
        input_symbol_hist.append(row["symbol"])
    port_on = pi['Portfolio Ongoing']
    i = 13
    port_on.iloc[2, 4] = 0
    port_on.iloc[6, 4] = 0
    port_on.iloc[7, 4] = 0
    try:
        while (str(port_on.iloc[i, 1]) != np.nan):
            for j in range(1, 15):
                port_on.iloc[i, j] = np.nan
            i = i + 1
    except:
        print("PI Portfolio ongoing Cleaned")
    output_writer_headers("M3T1PI.xlsx", port_on, 'Portfolio Ongoing', 1)

    port_s = pi['Portfolio Selected']
    i = 12
    try:
        while (str(port_s.iloc[i, 1]) != np.nan):
            for j in range(1, 17):
                port_s.iloc[i, j] = np.nan
            i = i + 1
    except:
        print("PI Portfolio Selected Cleaned")

    port_se = pi['Portfolio Selection']
    i = 6
    try:
        while (str(port_se.iloc[i, 1]) != np.nan):
            for j in range(1, 5):
                port_se.iloc[i, j] = np.nan
            i = i + 1
    except:
        print("PI Portfolio Selection Cleaned")

    df1 = pd.DataFrame(columns=port_s.columns)
    df1.drop(port_s.columns[1], inplace=True, axis=1)
    df1.insert(1, port_s.columns[1], input_symbol_hist)
    df2 = pd.DataFrame(columns=port_se.columns)
    df2.drop(port_se.columns[1], inplace=True, axis=1)
    df2.insert(1, port_se.columns[1], input_symbol_hist)
    port_s = pd.concat([port_s, df1])
    port_se = pd.concat([port_se, df2])
    output_writer_headers("M3T1PI.xlsx", port_se, 'Portfolio Selection', 1)
    output_writer_headers("M3T1PI.xlsx", port_s, 'Portfolio Selected', 1)
# In[3]:


#Historical data download
class historical_Data(EWrapper, EClient):
    
    def __init__(self):
        
        EClient.__init__(self, self)
        time.sleep(1)
        self.df = pd.DataFrame(columns = ["HistoricalData","Date","Open","High","Low","Close","Volume","Count", "WAP"])

#     def error(self, reqId, errorCode, errorString):
#         print("Error: ", reqId, " ", errorCode, " ", errorString)
    
    def historicalData(self, reqId, bar):
        df2 = pd.DataFrame({"HistoricalData":[ reqId], "Date":[ bar.date], "Open":[ bar.open],"High":[ bar.high], "Low":[ bar.low], "Close":[ bar.close], "Volume":[ bar.volume],"Count":[ bar.barCount], "WAP":[ bar.average]}) 
#         print("HistoricalData:", reqId, "Date:", bar.date, "Open:", bar.open,"High:", bar.high, "Low:", bar.low, "Close:", bar.close, "Volume:", bar.volume,"Count:", bar.barCount, "WAP:", bar.average)
        self.df = self.df.append(df2)
        
    def historicalDataEnd(self, reqId: int, start: str, end: str):
        print("historicalDataEnd sent!")
        self.stop()

    def stop(self):
        print("disconnecting!")
        self.done = True
        self.disconnect()
        time.sleep(1)
        
    def getData(self):
        return self.df
    
        
    def hist_data_download(i,symbol="EUR",secType="CASH",exchange="ISLAND",currency="USD",primaryExchange="NASDAQ",duration="1 D",bar_size="1 min",type_data="MIDPOINT"):
        app = historical_Data()
        app.connect("127.0.0.1", 7496, 0)
        time.sleep(1)
        contract = Contract()
        contract.symbol = symbol
        contract.secType = secType
        contract.exchange = exchange
        contract.currency = currency

        app.reqHistoricalData(i, contract, "", duration, bar_size, type_data, 0, 1, False,[])
        app.run()
        res = app.getData()
        res["Date"] = res["Date"].astype(str)       
        res["Date"] = res["Date"].apply(lambda x : parse(x).strftime("%m/%d/%Y"))
        output_writer("Complete_Output.xlsx",res,symbol,0)

        df1 = pd.DataFrame(list(zip(res["Date"], res["Close"])), columns =['Date', 'Close'])
        output_writer("M3T1DI.xlsx",df1,symbol,0)
        app.disconnect()
        
        
#place order
class order_place(EWrapper, EClient):
    
    def __init__(self):
        EClient.__init__(self, self)
        
    def error(self, reqId , errorCode, errorString):
         print("Error: ", reqId, " ", errorCode, " ", errorString)
            
    def nextValidId(self, orderId ):
        self.nextOrderId = orderId
        self.placeOrder(self.nextOrderId,self.contract, self.order)
        
    def orderStatus(self, orderId , status, filled, remaining, avgFillPrice,permId, parentId, lastFillPrice, clientId, whyHeld, mktCapPrice):
        print("OrderStatus. Id: ", orderId, ", Status: ", status, ", Filled:", filled, ", Remaining: ", remaining, ", LastFillPrice: ", lastFillPrice)
    
    def openOrder(self, orderId, contract, order, orderState):
        print("OpenOrder. ID:", orderId, contract.symbol, contract.secType,"@", contract.exchange, ":", order.action, order.orderType,order.totalQuantity, orderState.status)
    
    def execDetails(self, reqId, contract, execution):
        print("ExecDetails. ", reqId, contract.symbol, contract.secType,contract.currency, execution.execId,execution.orderId, execution.shares, execution.lastLiquidity)   

    def stop(self):
        self.done = True
        self.disconnect()
        
    def place_order(symbol,secType,currency,exchange,action,primaryExchange,totalQuantity,orderType,lmtPrice,stpPrice):
        app = order_place()
        app.nextOrderId = 0
        app.connect("127.0.0.1", 7496, 9)
        time.sleep(1)
        Timer(3, app.stop).start()
        app.contract = Contract()
        app.contract.symbol = symbol
        app.contract.secType = secType
        app.contract.exchange = exchange
        app.contract.currency = currency
        app.contract.primaryExchange = primaryExchange

        app.order = Order()
        app.order.action = action
        app.order.totalQuantity = totalQuantity
        app.order.orderType = orderType
        app.order.lmtPrice = lmtPrice
        if orderType == "STPLMT":
            app.order.auxPrice = stpPrice;
        app.run()
        app.disconnect()

        
class profile_data(EWrapper, EClient):
    def __init__(self):
        EClient.__init__(self, self)
        self.pf = pd.DataFrame(columns = ["Symbol","SecType","Exchange","Position","MarketPrice","MarketValue","AverageCost","UnrealizedPNL","RealizedPNL","AccountName"])
        self.netv = pd.DataFrame(columns = ["Key","Value","Currency","AccountName"])
                                 
    def error(self, reqId, errorCode, errorString):
        print("Error: ", reqId, " ", errorCode, " ", errorString)
        
    def nextValidId(self, orderId):
        self.start() 
        
    def updatePortfolio(self, contract: Contract, position:float, marketPrice: float, marketValue: float,averageCost: float, unrealizedPNL: float, realizedPNL: float, accountName: str):
#         print("UpdatePortfolio.", "Symbol:", contract.symbol, "SecType:", contract.secType, "Exchange:", contract.exchange,"Position:", position, "MarketPrice:", marketPrice, "MarketValue:", marketValue, "AverageCost:", averageCost,"UnrealizedPNL:", unrealizedPNL, "RealizedPNL:", realizedPNL, "AccountName:", accountName) 
        pf1 = pd.DataFrame({"Symbol":[contract.symbol],"SecType":[contract.secType],"Exchange":[contract.exchange],"Position":[position],"MarketPrice":[marketPrice],"MarketValue":[marketValue],"AverageCost":[averageCost],"UnrealizedPNL":[unrealizedPNL],"RealizedPNL":[realizedPNL],"AccountName":[accountName]})
        self.pf = self.pf.append(pf1)
        
    def updateAccountValue(self, key: str, val: str, currency: str, accountName: str):
#         print(" Key:", key, "Value:", val, "Currency:", currency, "AccountName:", accountName) 
        temp = pd.DataFrame({"Key":[key],"Value":[val],"Currency":[currency],"AccountName":[accountName]})
        self.netv = self.netv.append(temp)
#     def updateAccountTime(self, timeStamp: str): 
#         print("UpdateAccountTime. Time:", timeStamp) 
        
#     def accountDownloadEnd(self, accountName: str):
#         print("AccountDownloadEnd. Account:", accountName) 
        
    def start(self):
        # Account number can be omitted when using reqAccountUpdates with single account structure
        self.reqAccountUpdates(True, "") 
    
    def stop(self):
        self.reqAccountUpdates(False, "")
        self.done = True
        self.disconnect() 
        
    def getData(self):
        return self.pf,self.netv
    
    def profile_info():
        app = profile_data()
        app.nextOrderId = 0
        app.connect("127.0.0.1", 7496, 0) 
        Timer(5, app.stop).start() 
        app.run()
        init_output_file("m3t2OD1.xlsx",1)
        init_output_file("m3t2OD2.xlsx",1)
        
        res,res1=app.getData()
        output_writer("m3t2OD1.xlsx",res,"symbol",0)
        output_writer("m3t2OD2.xlsx",res1,"symbol",0)
        init_output_file("m3t2OD1.xlsx",2)
        init_output_file("m3t2OD2.xlsx",2)
        app.disconnect()


# In[4]:


import asyncio
import copy
import datetime
import logging
from contextlib import suppress
from typing import Awaitable, Iterator, List, Optional, Union

from eventkit import Event
import nest_asyncio
nest_asyncio.apply()
import ib_insync.util as util
from ib_insync.client import Client
from ib_insync.contract import Contract, ContractDescription, ContractDetails
from ib_insync.objects import (
    AccountValue, BarDataList, BarList, DepthMktDataDescription, Execution,
    ExecutionFilter, Fill, HistogramData, HistoricalNews, NewsArticle,
    NewsBulletin, NewsProvider, NewsTick, OptionChain, OptionComputation,
    PnL, PnLSingle, PortfolioItem, Position, PriceIncrement,
    RealTimeBarList, ScanDataList, ScannerSubscription, TagValue,
    TradeLogEntry)
from ib_insync.order import (
    BracketOrder, LimitOrder, Order, OrderState, OrderStatus, StopOrder, Trade)
from ib_insync.ticker import Ticker
from ib_insync.wrapper import Wrapper

__all__ = ['IB']


class IB:


    events = (
        'connectedEvent', 'disconnectedEvent', 'updateEvent',
        'pendingTickersEvent',
        'openAllOrderEvent', 'orderStatusEvent',
        'execDetailsEvent', 'commissionReportEvent', 'positionEvent', 'accountValueEvent',
        'accountSummaryEvent', 'pnlEvent', 'pnlSingleEvent',
        'scannerDataEvent', 'tickNewsEvent', 'newsBulletinEvent',
        'errorEvent', 'timeoutEvent')

    RequestTimeout = 0

    def __init__(self):
        self._createEvents()
        self.wrapper = Wrapper(self)
        self.client = Client(self.wrapper)
        self.client.apiEnd += self.disconnectedEvent
        self._logger = logging.getLogger('ib_insync.ib')

    def _createEvents(self):
        self.connectedEvent = Event('connectedEvent')
        self.disconnectedEvent = Event('disconnectedEvent')
        self.updateEvent = Event('updateEvent')
        self.pendingTickersEvent = Event('pendingTickersEvent')
        self.barUpdateEvent = Event('barUpdateEvent')
        self.newOrderEvent = Event('newOrderEvent')
        self.orderModifyEvent = Event('orderModifyEvent')
        self.cancelOrderEvent = Event('cancelOrderEvent')
        self.openAllOrderEvent = Event('openOrderEvent')
        self.orderStatusEvent = Event('orderStatusEvent')
        self.execDetailsEvent = Event('execDetailsEvent')
        self.commissionReportEvent = Event('commissionReportEvent')
        self.updatePortfolioEvent = Event('updatePortfolioEvent')
        self.positionEvent = Event('positionEvent')
        self.accountValueEvent = Event('accountValueEvent')
        self.accountSummaryEvent = Event('accountSummaryEvent')
        self.pnlEvent = Event('pnlEvent')
        self.pnlSingleEvent = Event('pnlSingleEvent')
        self.scannerDataEvent = Event('scannerDataEvent')
        self.tickNewsEvent = Event('tickNewsEvent')
        self.newsBulletinEvent = Event('newsBulletinEvent')
        self.errorEvent = Event('errorEvent')
        self.timeoutEvent = Event('timeoutEvent')

    def __del__(self):
        self.disconnect()

    def __enter__(self):
        return self

    def __exit__(self, *_exc):
        self.disconnect()

    def __repr__(self):
        conn = (f'connected to {self.client.host}:'
                f'{self.client.port} clientId={self.client.clientId}' if
                self.client.isConnected() else 'not connected')
        return f'<{self.__class__.__qualname__} {conn}>'

    def connect(
            self, host: str = '127.0.0.1', port: int = 7497, clientId: int = 1,
            timeout: float = 2, readonly: bool = False, account: str = ''):

        return self._run(self.connectAsync(
            host, port, clientId, timeout, readonly, account))


    def disconnect(self):

        if not self.client.isConnected():
            return
        stats = self.client.connectionStats()
        self._logger.info(
            f'Disconnecting from {self.client.host}:{self.client.port}, '
            f'{util.formatSI(stats.numBytesSent)}B sent '
            f'in {stats.numMsgSent} messages, '
            f'{util.formatSI(stats.numBytesRecv)}B received '
            f'in {stats.numMsgRecv} messages, '
            f'session time {util.formatSI(stats.duration)}s.')
        self.client.disconnect()


    def isConnected(self) -> bool:

        return self.client.isConnected()


    run = staticmethod(util.run)
    schedule = staticmethod(util.schedule)
    sleep = staticmethod(util.sleep)
    timeRange = staticmethod(util.timeRange)
    timeRangeAsync = staticmethod(util.timeRangeAsync)
    waitUntil = staticmethod(util.waitUntil)

    def _run(self, *awaitables: Awaitable):
        return util.run(*awaitables, timeout=self.RequestTimeout)

    def waitOnUpdate(self, timeout: float = 0) -> bool:

        if timeout:
            with suppress(asyncio.TimeoutError):
                util.run(asyncio.wait_for(self.updateEvent, timeout))
        else:
            util.run(self.updateEvent)
        return True


    def loopUntil(
            self, condition=None, timeout: float = 0) -> Iterator[object]:

        endTime = time.time() + timeout
        while True:
            test = condition and condition()
            if test:
                yield test
                return
            elif timeout and time.time() > endTime:
                yield False
                return
            else:
                yield test
            self.waitOnUpdate(endTime - time.time() if timeout else 0)


    def setTimeout(self, timeout: float = 60):

        self.wrapper.setTimeout(timeout)


    def managedAccounts(self) -> List[str]:

        return list(self.wrapper.accounts)


    def accountValues(self, account: str = '') -> List[AccountValue]:


        if account:
            return [v for v in self.wrapper.accountValues.values()
                    if v.account == account]
        else:
            return list(self.wrapper.accountValues.values())


    def accountSummary(self, account: str = '') -> List[AccountValue]:

        if not self.wrapper.acctSummary:
            # loaded on demand since it takes ca. 250 ms
            self.reqAccountSummary()
        if account:
            return [v for v in self.wrapper.acctSummary.values()
                    if v.account == account]
        else:
            return list(self.wrapper.acctSummary.values())


    def portfolio(self) -> List[PortfolioItem]:

        account = self.wrapper.accounts[0]
        return [v for v in self.wrapper.portfolio[account].values()]


    def positions(self, account: str = '') -> List[Position]:

        if account:
            return list(self.wrapper.positions[account].values())
        else:
            return [v for d in self.wrapper.positions.values()
                    for v in d.values()]


    def pnl(self, account='', modelCode='') -> List[PnL]:

        return [v for v in self.wrapper.reqId2PnL.values() if
                (not account or v.account == account)
                and (not modelCode or v.modelCode == modelCode)]


    def pnlSingle(
            self, account: str = '', modelCode: str = '',
            conId: int = 0) -> List[PnLSingle]:

        return [v for v in self.wrapper.reqId2PnlSingle.values() if
                (not account or v.account == account)
                and (not modelCode or v.modelCode == modelCode)
                and (not conId or v.conId == conId)]


    def trades(self) -> List[Trade]:

        return list(self.wrapper.trades.values())


    def openTrades(self) -> List[Trade]:

        return [v for v in self.wrapper.trades.values()
                if v.orderStatus.status not in OrderStatus.DoneStates]


    def orders(self) -> List[Order]:

        return list(
            trade.order for trade in self.wrapper.trades.values())


    def openOrders(self) -> List[Order]:

        return [trade.order for trade in self.wrapper.trades.values()
                if trade.orderStatus.status not in OrderStatus.DoneStates]


    def fills(self) -> List[Fill]:

        return list(self.wrapper.fills.values())


    def executions(self) -> List[Execution]:

        return list(fill.execution for fill in self.wrapper.fills.values())


    def ticker(self, contract: Contract) -> Ticker:

        return self.wrapper.tickers.get(id(contract))


    def tickers(self) -> List[Ticker]:

        return list(self.wrapper.tickers.values())


    def pendingTickers(self) -> List[Ticker]:

        return list(self.wrapper.pendingTickers)


    def realtimeBars(self) -> BarList:

        return BarList(self.wrapper.reqId2Subscriber.values())


    def newsTicks(self) -> List[NewsTick]:

        return self.wrapper.newsTicks


    def newsBulletins(self) -> List[NewsBulletin]:

        return list(self.wrapper.msgId2NewsBulletin.values())


    def reqTickers(
            self, *contracts: Contract,
            regulatorySnapshot: bool = False) -> List[Ticker]:

        return self._run(
            self.reqTickersAsync(
                *contracts, regulatorySnapshot=regulatorySnapshot))


    def qualifyContracts(self, *contracts: Contract) -> List[Contract]:

        return self._run(self.qualifyContractsAsync(*contracts))


    def bracketOrder(
            self, action: str, quantity: float,
            limitPrice: float, takeProfitPrice: float,
            stopLossPrice: float, **kwargs) -> BracketOrder:

        assert action in ('BUY', 'SELL')
        reverseAction = 'BUY' if action == 'SELL' else 'SELL'
        parent = LimitOrder(
            action, quantity, limitPrice,
            orderId=self.client.getReqId(),
            transmit=False,
            **kwargs)
        takeProfit = LimitOrder(
            reverseAction, quantity, takeProfitPrice,
            orderId=self.client.getReqId(),
            transmit=False,
            parentId=parent.orderId,
            **kwargs)
        stopLoss = StopOrder(
            reverseAction, quantity, stopLossPrice,
            orderId=self.client.getReqId(),
            transmit=True,
            parentId=parent.orderId,
            **kwargs)
        return BracketOrder(parent, takeProfit, stopLoss)


    @staticmethod
    def oneCancelsAll(
            orders: List[Order], ocaGroup: str, ocaType: int) -> List[Order]:

        for o in orders:
            o.ocaGroup = ocaGroup
            o.ocaType = ocaType
        return orders


    def whatIfOrder(self, contract: Contract, order: Order) -> OrderState:

        return self._run(self.whatIfOrderAsync(contract, order))


    def placeOrder(self, contract: Contract, order: Order) -> Trade:

        orderId = order.orderId or self.client.getReqId()
        self.client.placeOrder(orderId, contract, order)
        now = datetime.datetime.now(datetime.timezone.utc)
        key = self.wrapper.orderKey(
            self.wrapper.clientId, orderId, order.permId)
        trade = self.wrapper.trades.get(key)
        if trade:
            # this is a modification of an existing order
            assert trade.orderStatus.status not in OrderStatus.DoneStates
            logEntry = TradeLogEntry(now, trade.orderStatus.status, 'Modify')
            trade.log.append(logEntry)
            self._logger.info(f'placeOrder: Modify order {trade}')
            trade.modifyEvent.emit(trade)
            self.orderModifyEvent.emit(trade)
        else:
            # this is a new order
            order.clientId = self.wrapper.clientId
            order.orderId = orderId
            orderStatus = OrderStatus(
                orderId=orderId, status=OrderStatus.PendingSubmit)
            logEntry = TradeLogEntry(now, orderStatus.status, '')
            trade = Trade(
                contract, order, orderStatus, [], [logEntry])
            self.wrapper.trades[key] = trade
            self._logger.info(f'placeOrder: New order {trade}')
            self.newOrderEvent.emit(trade)
        return trade


    def cancelOrder(self, order: Order) -> Trade:

        self.client.cancelOrder(order.orderId)
        now = datetime.datetime.now(datetime.timezone.utc)
        key = self.wrapper.orderKey(
            order.clientId, order.orderId, order.permId)
        trade = self.wrapper.trades.get(key)
        if trade:
            if not trade.isDone():
                status = trade.orderStatus.status
                if (status == OrderStatus.PendingSubmit and not order.transmit
                        or status == OrderStatus.Inactive):
                    newStatus = OrderStatus.Cancelled
                else:
                    newStatus = OrderStatus.PendingCancel
                logEntry = TradeLogEntry(now, newStatus, '')
                trade.log.append(logEntry)
                trade.orderStatus.status = newStatus
                self._logger.info(f'cancelOrder: {trade}')
                trade.cancelEvent.emit(trade)
                trade.statusEvent.emit(trade)
                self.cancelOrderEvent.emit(trade)
                self.orderStatusEvent.emit(trade)
                if newStatus == OrderStatus.Cancelled:
                    trade.cancelledEvent.emit(trade)
        else:
            self._logger.error(f'cancelOrder: Unknown orderId {order.orderId}')
        return trade


    def reqGlobalCancel(self):

        self.client.reqGlobalCancel()
        self._logger.info(f'reqGlobalCancel')


    def reqCurrentTime(self) -> datetime.datetime:

        return self._run(self.reqCurrentTimeAsync())


    def reqAccountUpdates(self, account: str = ''):

        self._run(self.reqAccountUpdatesAsync(account))


    def reqAccountUpdatesMulti(
            self, account: str = '', modelCode: str = ''):

        self._run(self.reqAccountUpdatesMultiAsync(account, modelCode))


    def reqAccountSummary(self):

        self._run(self.reqAccountSummaryAsync())


    def reqAutoOpenOrders(self, autoBind: bool = True):

        self.client.reqAutoOpenOrders(autoBind)


    def reqOpenOrders(self) -> List[Order]:

        return self._run(self.reqOpenOrdersAsync())


    def reqAllOpenOrders(self) -> List[Order]:

        return self._run(self.reqAllOpenOrdersAsync())


    def reqCompletedOrders(self, apiOnly: bool) -> List[Trade]:

        return self._run(self.reqCompletedOrdersAsync(apiOnly))


    def reqExecutions(
            self, execFilter: ExecutionFilter = None) -> List[Fill]:

        return self._run(self.reqExecutionsAsync(execFilter))


    def reqPositions(self) -> List[Position]:

        return self._run(self.reqPositionsAsync())


    def reqPnL(self, account: str, modelCode: str = '') -> PnL:

        key = (account, modelCode)
        assert key not in self.wrapper.pnlKey2ReqId
        reqId = self.client.getReqId()
        self.wrapper.pnlKey2ReqId[key] = reqId
        pnl = PnL(account, modelCode)
        self.wrapper.reqId2PnL[reqId] = pnl
        self.client.reqPnL(reqId, account, modelCode)
        return pnl


    def cancelPnL(self, account, modelCode: str = ''):

        key = (account, modelCode)
        reqId = self.wrapper.pnlKey2ReqId.pop(key, None)
        if reqId:
            self.client.cancelPnL(reqId)
            self.wrapper.reqId2PnL.pop(reqId, None)
        else:
            self._logger.error(
                'cancelPnL: No subscription for '
                f'account {account}, modelCode {modelCode}')


    def reqPnLSingle(
            self, account: str, modelCode: str, conId: int) -> PnLSingle:

        key = (account, modelCode, conId)
        assert key not in self.wrapper.pnlSingleKey2ReqId
        reqId = self.client.getReqId()
        self.wrapper.pnlSingleKey2ReqId[key] = reqId
        pnlSingle = PnLSingle(account, modelCode, conId)
        self.wrapper.reqId2PnlSingle[reqId] = pnlSingle
        self.client.reqPnLSingle(reqId, account, modelCode, conId)
        return pnlSingle


    def cancelPnLSingle(
            self, account: str, modelCode: str, conId: int):

        key = (account, modelCode, conId)
        reqId = self.wrapper.pnlSingleKey2ReqId.pop(key, None)
        if reqId:
            self.client.cancelPnLSingle(reqId)
            self.wrapper.reqId2PnlSingle.pop(reqId, None)
        else:
            self._logger.error(
                'cancelPnLSingle: No subscription for '
                f'account {account}, modelCode {modelCode}, conId {conId}')


    def reqContractDetails(self, contract: Contract) -> List[ContractDetails]:

        return self._run(self.reqContractDetailsAsync(contract))


    def reqMatchingSymbols(self, pattern: str) -> List[ContractDescription]:

        return self._run(self.reqMatchingSymbolsAsync(pattern))


    def reqMarketRule(self, marketRuleId: int) -> PriceIncrement:

        return self._run(self.reqMarketRuleAsync(marketRuleId))


    def reqRealTimeBars(
            self, contract: Contract, barSize: int,
            whatToShow: str, useRTH: bool,
            realTimeBarsOptions: List[TagValue] = []) -> RealTimeBarList:

        reqId = self.client.getReqId()
        bars = RealTimeBarList()
        bars.reqId = reqId
        bars.contract = contract
        bars.barSize = barSize
        bars.whatToShow = whatToShow
        bars.useRTH = useRTH
        bars.realTimeBarsOptions = realTimeBarsOptions or []
        self.wrapper.startSubscription(reqId, bars, contract)
        self.client.reqRealTimeBars(
            reqId, contract, barSize, whatToShow, useRTH, realTimeBarsOptions)
        return bars


    def cancelRealTimeBars(self, bars: RealTimeBarList):

        self.client.cancelRealTimeBars(bars.reqId)
        self.wrapper.endSubscription(bars)


    def reqHistoricalData(
            self, contract: Contract,
            endDateTime: Union[datetime.datetime, datetime.date, str, None],
            durationStr: str, barSizeSetting: str, whatToShow: str,
            useRTH: bool, formatDate: int = 1, keepUpToDate: bool = False,
            chartOptions: List[TagValue] = []) -> BarDataList:

        return self._run(
            self.reqHistoricalDataAsync(
                contract, endDateTime, durationStr, barSizeSetting, whatToShow,
                useRTH, formatDate, keepUpToDate, chartOptions))


    def cancelHistoricalData(self, bars: BarDataList):

        self.client.cancelHistoricalData(bars.reqId)
        self.wrapper.endSubscription(bars)


    def reqHistoricalTicks(
            self, contract: Contract,
            startDateTime: Union[str, datetime.date],
            endDateTime: Union[str, datetime.date],
            numberOfTicks: int, whatToShow: str, useRth: bool,
            ignoreSize: bool = False,
            miscOptions: List[TagValue] = []) -> List:

        return self._run(
            self.reqHistoricalTicksAsync(
                contract, startDateTime, endDateTime, numberOfTicks,
                whatToShow, useRth, ignoreSize, miscOptions))


    def reqMarketDataType(self, marketDataType: int):

        self.client.reqMarketDataType(marketDataType)


    def reqHeadTimeStamp(
            self, contract: Contract, whatToShow: str,
            useRTH: bool, formatDate: int = 1) -> datetime.datetime:

        return self._run(
            self.reqHeadTimeStampAsync(
                contract, whatToShow, useRTH, formatDate))


    def reqMktData(
            self, contract: Contract, genericTickList: str = '',
            snapshot: bool = False, regulatorySnapshot: bool = False,
            mktDataOptions: List[TagValue] = None) -> Ticker:

        reqId = self.client.getReqId()
        ticker = self.wrapper.startTicker(reqId, contract, 'mktData')
        self.client.reqMktData(
            reqId, contract, genericTickList, snapshot,
            regulatorySnapshot, mktDataOptions)
        return ticker


    def cancelMktData(self, contract: Contract):

        ticker = self.ticker(contract)
        reqId = self.wrapper.endTicker(ticker, 'mktData')
        if reqId:
            self.client.cancelMktData(reqId)
        else:
            self._logger.error(
                'cancelMktData: ' f'No reqId found for contract {contract}')


    def reqTickByTickData(
            self, contract: Contract, tickType: str,
            numberOfTicks: int = 0, ignoreSize: bool = False) -> Ticker:

        reqId = self.client.getReqId()
        ticker = self.wrapper.startTicker(reqId, contract, tickType)
        self.client.reqTickByTickData(
            reqId, contract, tickType, numberOfTicks, ignoreSize)
        return ticker


    def cancelTickByTickData(self, contract: Contract, tickType: str):

        ticker = self.ticker(contract)
        reqId = self.wrapper.endTicker(ticker, tickType)
        if reqId:
            self.client.cancelTickByTickData(reqId)
        else:
            self._logger.error(
                f'cancelMktData: No reqId found for contract {contract}')


    def reqMktDepthExchanges(self) -> List[DepthMktDataDescription]:

        return self._run(self.reqMktDepthExchangesAsync())


    def reqMktDepth(
            self, contract: Contract, numRows: int = 5,
            isSmartDepth: bool = False, mktDepthOptions=None) -> Ticker:

        reqId = self.client.getReqId()
        ticker = self.wrapper.startTicker(reqId, contract, 'mktDepth')
        self.client.reqMktDepth(
            reqId, contract, numRows, isSmartDepth, mktDepthOptions)
        return ticker


    def cancelMktDepth(self, contract: Contract, isSmartDepth=False):

        ticker = self.ticker(contract)
        reqId = self.wrapper.endTicker(ticker, 'mktDepth')
        if reqId:
            self.client.cancelMktDepth(reqId, isSmartDepth)
        else:
            self._logger.error(
                f'cancelMktDepth: No reqId found for contract {contract}')


    def reqHistogramData(
            self, contract: Contract,
            useRTH: bool, period: str) -> List[HistogramData]:

        return self._run(
            self.reqHistogramDataAsync(contract, useRTH, period))


    def reqFundamentalData(
            self, contract: Contract, reportType: str,
            fundamentalDataOptions: List[TagValue] = []) -> str:

        return self._run(
            self.reqFundamentalDataAsync(
                contract, reportType, fundamentalDataOptions))


    def reqScannerData(
            self, subscription: ScannerSubscription,
            scannerSubscriptionOptions: List[TagValue] = [],
            scannerSubscriptionFilterOptions: List[TagValue] = []) -> \
            ScanDataList:

        return self._run(
            self.reqScannerDataAsync(
                subscription, scannerSubscriptionOptions,
                scannerSubscriptionFilterOptions))


    def reqScannerSubscription(
            self, subscription: ScannerSubscription,
            scannerSubscriptionOptions: List[TagValue] = [],
            scannerSubscriptionFilterOptions:
            List[TagValue] = []) -> ScanDataList:

        reqId = self.client.getReqId()
        dataList = ScanDataList()
        dataList.reqId = reqId
        dataList.subscription = subscription
        dataList.scannerSubscriptionOptions = scannerSubscriptionOptions or []
        dataList.scannerSubscriptionFilterOptions =             scannerSubscriptionFilterOptions or []
        self.wrapper.startSubscription(reqId, dataList)
        self.client.reqScannerSubscription(
            reqId, subscription, scannerSubscriptionOptions,
            scannerSubscriptionFilterOptions)
        return dataList


    def cancelScannerSubscription(self, dataList: ScanDataList):

        self.client.cancelScannerSubscription(dataList.reqId)
        self.wrapper.endSubscription(dataList)


    def reqScannerParameters(self) -> str:

        return self._run(self.reqScannerParametersAsync())


    def calculateImpliedVolatility(
            self, contract: Contract,
            optionPrice: float, underPrice: float,
            implVolOptions: List[TagValue] = []) -> OptionComputation:

        return self._run(
            self.calculateImpliedVolatilityAsync(
                contract, optionPrice, underPrice, implVolOptions))


    def calculateOptionPrice(
            self, contract: Contract,
            volatility: float, underPrice: float,
            optPrcOptions=None) -> OptionComputation:

        return self._run(
            self.calculateOptionPriceAsync(
                contract, volatility, underPrice, optPrcOptions))


    def reqSecDefOptParams(
            self, underlyingSymbol: str,
            futFopExchange: str, underlyingSecType: str,
            underlyingConId: int) -> List[OptionChain]:

        return self._run(
            self.reqSecDefOptParamsAsync(
                underlyingSymbol, futFopExchange,
                underlyingSecType, underlyingConId))


    def exerciseOptions(
            self, contract: Contract, exerciseAction: int,
            exerciseQuantity: int, account: str, override: int):

        reqId = self.client.getReqId()
        self.client.exerciseOptions(
            reqId, contract, exerciseAction, exerciseQuantity,
            account, override)


    def reqNewsProviders(self) -> List[NewsProvider]:

        return self._run(self.reqNewsProvidersAsync())


    def reqNewsArticle(
            self, providerCode: str, articleId: str,
            newsArticleOptions: List[TagValue] = None) -> NewsArticle:

        return self._run(
            self.reqNewsArticleAsync(
                providerCode, articleId, newsArticleOptions))


    def reqHistoricalNews(
            self, conId: int, providerCodes: str,
            startDateTime: Union[str, datetime.date],
            endDateTime: Union[str, datetime.date],
            totalResults: int,
            historicalNewsOptions: List[TagValue] = None) -> HistoricalNews:

        return self._run(
            self.reqHistoricalNewsAsync(
                conId, providerCodes, startDateTime, endDateTime,
                totalResults, historicalNewsOptions))


    def reqNewsBulletins(self, allMessages: bool):

        self.client.reqNewsBulletins(allMessages)


    def cancelNewsBulletins(self):

        self.client.cancelNewsBulletins()


    def requestFA(self, faDataType: int):

        return self._run(self.requestFAAsync(faDataType))


    def replaceFA(self, faDataType: int, xml: str):

        self.client.replaceFA(faDataType, xml)


    # now entering the parallel async universe

    async def connectAsync(
            self, host: str = '127.0.0.1', port: int = 7497,
            clientId: int = 1, timeout: float = 2, readonly: bool = False,
            account: str = ''):

        async def connect():
            self.wrapper.clientId = clientId
            await self.client.connectAsync(host, port, clientId, timeout)
            if not readonly and self.client.serverVersion() >= 150:
                await self.reqCompletedOrdersAsync(False)
            accounts = self.client.getAccounts()
            await asyncio.gather(
                self.reqAccountUpdatesAsync(account or accounts[0]),
                *(self.reqAccountUpdatesMultiAsync(a) for a in accounts),
                self.reqPositionsAsync(),
                self.reqExecutionsAsync())
            if clientId == 0:
                # autobind manual orders
                self.reqAutoOpenOrders(True)
            self._logger.info('Synchronization complete')
            self.connectedEvent.emit()

        if not self.isConnected():
            try:
                await asyncio.wait_for(connect(), timeout or None)
            except Exception:
                self.disconnect()
                raise
        else:
            self._logger.warn('Already connected')
        return self


    async def qualifyContractsAsync(self, *contracts: Contract) ->             List[Contract]:
        detailsLists = await asyncio.gather(
            *(self.reqContractDetailsAsync(c) for c in contracts))
        result = []
        for contract, detailsList in zip(contracts, detailsLists):
            if not detailsList:
                self._logger.error(f'Unknown contract: {contract}')
            elif len(detailsList) > 1:
                possibles = [details.contract for details in detailsList]
                self._logger.error(
                    f'Ambiguous contract: {contract}, '
                    f'possibles are {possibles}')
            else:
                c = detailsList[0].contract
                expiry = c.lastTradeDateOrContractMonth
                if expiry:
                    # remove time and timezone part as it will cause problems
                    expiry = expiry.split()[0]
                    c.lastTradeDateOrContractMonth = expiry
                if contract.exchange == 'SMART':
                    # overwriting 'SMART' exchange can create invalid contract
                    c.exchange = contract.exchange
                util.dataclassUpdate(contract, c)
                result.append(contract)
        return result


    async def reqTickersAsync(
            self, *contracts: Contract, regulatorySnapshot: bool = False) -> \
            List[Ticker]:
        futures = []
        tickers = []
        for contract in contracts:
            reqId = self.client.getReqId()
            future = self.wrapper.startReq(reqId, contract)
            futures.append(future)
            ticker = self.wrapper.startTicker(reqId, contract, 'snapshot')
            tickers.append(ticker)
            self.client.reqMktData(
                reqId, contract, '', True, regulatorySnapshot, [])
        await asyncio.gather(*futures)
        for ticker in tickers:
            self.wrapper.endTicker(ticker, 'snapshot')
        return tickers


    def whatIfOrderAsync(self, contract: Contract, order: Order) ->             Awaitable[OrderState]:
        whatIfOrder = copy.copy(order)
        whatIfOrder.whatIf = True
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId, contract)
        self.client.placeOrder(reqId, contract, whatIfOrder)
        return future


    def reqCurrentTimeAsync(self) -> Awaitable[datetime.datetime]:
        future = self.wrapper.startReq('currentTime')
        self.client.reqCurrentTime()
        return future


    def reqAccountUpdatesAsync(self, account: str) -> Awaitable[None]:
        future = self.wrapper.startReq('accountValues')
        self.client.reqAccountUpdates(True, account)
        return future


    def reqAccountUpdatesMultiAsync(
            self, account: str, modelCode: str = '') -> Awaitable[None]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId)
        self.client.reqAccountUpdatesMulti(reqId, account, modelCode, False)
        return future


    def reqAccountSummaryAsync(self) -> Awaitable[None]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId)
        tags = (
            'AccountType,NetLiquidation,TotalCashValue,SettledCash,'
            'AccruedCash,BuyingPower,EquityWithLoanValue,'
            'PreviousEquityWithLoanValue,GrossPositionValue,ReqTEquity,'
            'ReqTMargin,SMA,InitMarginReq,MaintMarginReq,AvailableFunds,'
            'ExcessLiquidity,Cushion,FullInitMarginReq,FullMaintMarginReq,'
            'FullAvailableFunds,FullExcessLiquidity,LookAheadNextChange,'
            'LookAheadInitMarginReq,LookAheadMaintMarginReq,'
            'LookAheadAvailableFunds,LookAheadExcessLiquidity,'
            'HighestSeverity,DayTradesRemaining,Leverage,$LEDGER:ALL')
        self.client.reqAccountSummary(reqId, 'All', tags)
        return future


    def reqOpenOrdersAsync(self) -> Awaitable[List[Order]]:
        future = self.wrapper.startReq('openOrders')
        self.client.reqOpenOrders()
        return future


    def reqAllOpenOrdersAsync(self) -> Awaitable[List[Order]]:
        future = self.wrapper.startReq('openOrders')
        self.client.reqAllOpenOrders()
        return future


    def reqCompletedOrdersAsync(self, apiOnly: bool) -> Awaitable[List[Trade]]:
        future = self.wrapper.startReq('completedOrders')
        self.client.reqCompletedOrders(apiOnly)
        return future


    def reqExecutionsAsync(
            self, execFilter: ExecutionFilter = None) -> Awaitable[List[Fill]]:
        execFilter = execFilter or ExecutionFilter()
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId)
        self.client.reqExecutions(reqId, execFilter)
        return future


    def reqPositionsAsync(self) -> Awaitable[List[Position]]:
        future = self.wrapper.startReq('positions')
        self.client.reqPositions()
        return future


    def reqContractDetailsAsync(self, contract: Contract) ->             Awaitable[List[ContractDetails]]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId, contract)
        self.client.reqContractDetails(reqId, contract)
        return future


    async def reqMatchingSymbolsAsync(self, pattern: str) ->             Optional[List[ContractDescription]]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId)
        self.client.reqMatchingSymbols(reqId, pattern)
        try:
            await asyncio.wait_for(future, 4)
            return future.result()
        except asyncio.TimeoutError:
            self._logger.error('reqMatchingSymbolsAsync: Timeout')
            return None


    async def reqMarketRuleAsync(
            self, marketRuleId: int) -> Optional[PriceIncrement]:
        future = self.wrapper.startReq(f'marketRule-{marketRuleId}')
        try:
            self.client.reqMarketRule(marketRuleId)
            await asyncio.wait_for(future, 1)
            return future.result()
        except asyncio.TimeoutError:
            self._logger.error('reqMarketRuleAsync: Timeout')
            return None


    def reqHistoricalDataAsync(
            self, contract: Contract,
            endDateTime: Union[datetime.datetime, datetime.date, str, None],
            durationStr: str, barSizeSetting: str,
            whatToShow: str, useRTH: bool,
            formatDate: int = 1, keepUpToDate: bool = False,
            chartOptions: List[TagValue] = []) -> \
            Awaitable[BarDataList]:
        reqId = self.client.getReqId()
        bars = BarDataList()
        bars.reqId = reqId
        bars.contract = contract
        bars.endDateTime = endDateTime
        bars.durationStr = durationStr
        bars.barSizeSetting = barSizeSetting
        bars.whatToShow = whatToShow
        bars.useRTH = useRTH
        bars.formatDate = formatDate
        bars.keepUpToDate = keepUpToDate
        bars.chartOptions = chartOptions or []
        future = self.wrapper.startReq(reqId, contract, container=bars)
        if keepUpToDate:
            self.wrapper.startSubscription(reqId, bars, contract)
        end = util.formatIBDatetime(endDateTime)
        self.client.reqHistoricalData(
            reqId, contract, end, durationStr, barSizeSetting,
            whatToShow, useRTH, formatDate, keepUpToDate, chartOptions)
        return future


    def reqHistoricalTicksAsync(
            self, contract: Contract,
            startDateTime: Union[str, datetime.date],
            endDateTime: Union[str, datetime.date],
            numberOfTicks: int, whatToShow: str, useRth: bool,
            ignoreSize: bool = False,
            miscOptions: List[TagValue] = []) -> Awaitable[List]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId, contract)
        start = util.formatIBDatetime(startDateTime)
        end = util.formatIBDatetime(endDateTime)
        self.client.reqHistoricalTicks(
            reqId, contract, start, end, numberOfTicks, whatToShow, useRth,
            ignoreSize, miscOptions)
        return future


    def reqHeadTimeStampAsync(
            self, contract: Contract, whatToShow: str,
            useRTH: bool, formatDate: int) -> Awaitable[datetime.datetime]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId, contract)
        self.client.reqHeadTimeStamp(
            reqId, contract, whatToShow, useRTH, formatDate)
        return future


    def reqMktDepthExchangesAsync(self) ->             Awaitable[List[DepthMktDataDescription]]:
        future = self.wrapper.startReq('mktDepthExchanges')
        self.client.reqMktDepthExchanges()
        return future


    def reqHistogramDataAsync(
            self, contract: Contract, useRTH: bool, period: str) -> \
            Awaitable[List[HistogramData]]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId, contract)
        self.client.reqHistogramData(reqId, contract, useRTH, period)
        return future


    def reqFundamentalDataAsync(
            self, contract: Contract, reportType: str,
            fundamentalDataOptions: List[TagValue] = []) -> \
            Awaitable[str]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId, contract)
        self.client.reqFundamentalData(
            reqId, contract, reportType, fundamentalDataOptions)
        return future


    async def reqScannerDataAsync(
            self, subscription: ScannerSubscription,
            scannerSubscriptionOptions: List[TagValue] = [],
            scannerSubscriptionFilterOptions: List[TagValue] = []) \
            -> ScanDataList:
        dataList = self.reqScannerSubscription(
            subscription, scannerSubscriptionOptions or [],
            scannerSubscriptionFilterOptions or [])
        future = self.wrapper.startReq(dataList.reqId, container=dataList)
        await future
        self.client.cancelScannerSubscription(dataList.reqId)
        return future.result()


    def reqScannerParametersAsync(self) -> Awaitable[str]:
        future = self.wrapper.startReq('scannerParams')
        self.client.reqScannerParameters()
        return future


    async def calculateImpliedVolatilityAsync(
            self, contract: Contract,
            optionPrice: float, underPrice: float,
            implVolOptions: List[TagValue] = []) -> \
            Optional[OptionComputation]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId, contract)
        self.client.calculateImpliedVolatility(
            reqId, contract, optionPrice, underPrice, implVolOptions)
        try:
            await asyncio.wait_for(future, 4)
            return future.result()
        except asyncio.TimeoutError:
            self._logger.error('calculateImpliedVolatilityAsync: Timeout')
            return None
        finally:
            self.client.cancelCalculateImpliedVolatility(reqId)


    async def calculateOptionPriceAsync(
            self, contract: Contract,
            volatility: float, underPrice: float,
            optPrcOptions: List[TagValue] = []) -> Optional[OptionComputation]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId, contract)
        self.client.calculateOptionPrice(
            reqId, contract, volatility, underPrice, optPrcOptions)
        try:
            await asyncio.wait_for(future, 4)
            return future.result()
        except asyncio.TimeoutError:
            self._logger.error('calculateOptionPriceAsync: Timeout')
            return None
        finally:
            self.client.cancelCalculateOptionPrice(reqId)


    def reqSecDefOptParamsAsync(
            self, underlyingSymbol: str,
            futFopExchange: str, underlyingSecType: str,
            underlyingConId: int) -> Awaitable[List[OptionChain]]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId)
        self.client.reqSecDefOptParams(
            reqId, underlyingSymbol, futFopExchange,
            underlyingSecType, underlyingConId)
        return future


    def reqNewsProvidersAsync(self) -> Awaitable[List[NewsProvider]]:
        future = self.wrapper.startReq('newsProviders')
        self.client.reqNewsProviders()
        return future


    def reqNewsArticleAsync(
            self, providerCode: str, articleId: str,
            newsArticleOptions: Optional[List[TagValue]]) -> \
            Awaitable[NewsArticle]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId)
        self.client.reqNewsArticle(
            reqId, providerCode, articleId, newsArticleOptions)
        return future


    async def reqHistoricalNewsAsync(
            self, conId: int, providerCodes: str,
            startDateTime: Union[str, datetime.date],
            endDateTime: Union[str, datetime.date],
            totalResults: int,
            historicalNewsOptions: List[TagValue] = None) -> \
            Optional[HistoricalNews]:
        reqId = self.client.getReqId()
        future = self.wrapper.startReq(reqId)
        start = util.formatIBDatetime(startDateTime)
        end = util.formatIBDatetime(endDateTime)
        self.client.reqHistoricalNews(
            reqId, conId, providerCodes, start, end,
            totalResults, historicalNewsOptions)
        try:
            await asyncio.wait_for(future, 4)
            return future.result()
        except asyncio.TimeoutError:
            self._logger.error('reqHistoricalNewsAsync: Timeout')
            return None


    async def requestFAAsync(self, faDataType: int):
        future = self.wrapper.startReq('requestFA')
        self.client.requestFA(faDataType)
        try:
            await asyncio.wait_for(future, 4)
            return future.result()
        except asyncio.TimeoutError:
            self._logger.error('requestFAAsync: Timeout')


def placed_order_history():
    asyncio.get_event_loop().set_debug(True)
    util.logToConsole(logging.DEBUG)
    ib = IB()
    
    try:
        ib.connect('127.0.0.1', 7496, clientId=str(time.time()).split(".")[0] )
        output=ib.reqAllOpenOrders()
    except:
        ib.disconnect()
    ib.disconnect()
    out_df=pd.DataFrame(columns=['orderId','permId','action','totalQuantity','orderType','lmtPrice','auxPrice','tif','ocaType','rule80A','trailStopPrice','openClose','eTradeOnly','firmQuoteOnly','volatilityType','deltaNeutralOrderType','referencePriceType','account','clearingIntent','adjustedOrderType','cashQty','dontUseAutoPriceForHedge'])
    for i in range(0,len(output)):
        temp=str(output[i]).replace('(',',').replace('=',',').replace(')',',').replace(' ','').replace("'",'').split(",")
        del temp[0]
        del temp[-1]
        temp_df=pd.DataFrame([temp[1::2]],columns=temp[::2])
#         print(temp_df)
        out_df=out_df.append(temp_df)
    init_output_file("m3t2OD3.xlsx",1)
    output_writer("m3t2OD3.xlsx",out_df,"symbol",0)
    init_output_file("m3t2OD3.xlsx",2)


# In[5]:


def data_download():
    global df
    path = 'Historical_data_inputs.xlsx'
    df = pd.read_excel(path, sheet_name="Sheet1", encoding = 'unicode_escape')


    init_output_file("Complete_Output.xlsx",1)
    init_output_file("M3T1DI.xlsx",1)


    for i in range(df.shape[0]):
        try:
            historical_Data.hist_data_download(i,df.iloc[i]["symbol"],df.iloc[i]["secType"],df.iloc[i]["exchange"],df.iloc[i]["currency"],df.iloc[i]["primaryExchange"],df.iloc[i]["duration"],df.iloc[i]["bar_size"],df.iloc[i]["type_data"])
            print("Row  " ,i+1,"  Completed")
        except Exception as e:
            print("problem with :",i+1,e)

    init_output_file("Complete_Output.xlsx",2)
    init_output_file("M3T1DI.xlsx",2)
    print("Data Downloaded")
    
def order_placer():
    path = 'new_place_order.xlsx'
    df = pd.read_excel(path, sheet_name="New_place_order", encoding = 'unicode_escape')
    path = 'M3T1PI.xlsx'
    pi = pd.read_excel(path, sheet_name=None, encoding='unicode_escape', header=None)
    on_pi = pi['Portfolio Ongoing']
    max_orders = int(float(on_pi.iloc[3,11]))
    print(max_orders)
    for i in range(df.shape[0]):
        if max_orders > 0:
            max_orders = max_orders - 1
            try:
                time.sleep(1)
                if df.action[i] == "BUY" or df.action[i] == "SELL":
                    order_place.place_order(df.symbol[i],df.secType[i],df.currency[i],df.exchange[i],df.action[i],df.primaryExchange[i],df.totalQuantity[i],df.orderType[i],df.lmtPrice[i],df.stpPrice[i])
            except Exception as e:
                print("Error in order",e)
        else:
            break
    print("Order Placed")

def stoploss_order_placer():
    path = 'stoploss_place_order.xlsx'
    df = pd.read_excel(path, sheet_name="stoploss_place_order", encoding = 'unicode_escape')
    for i in range(df.shape[0]):
        try:
            time.sleep(1)
            if df.action[i] == "BUY" or df.action[i] == "SELL":
                order_place.place_order(df.symbol[i],df.secType[i],df.currency[i],df.exchange[i],df.action[i],df.primaryExchange[i],df.totalQuantity[i],df.orderType[i],df.lmtPrice[i],df.stpPrice[i])
        except Exception as e:
            print("Error in order",e)
    print("Order Placed")
# In[6]:


class create_fr():

    def extract_index(self,date_, relative_, dataset):
        date_ = date_.date()
        year = date_.year
        month = date_.month
        day = date_.day
        relative_date = date(year, month, day)-relativedelta(months=+relative_)
        try:
            return dataset[dataset['Date'].isin([relative_date])].index[0]
        except:
            for i in range(1,30):
                if relative_date-relativedelta(days=i) in list(dataset['Date']):
                    relative_date = relative_date-relativedelta(days=i)
                    return dataset[dataset['Date'].isin([relative_date])].index[0]
                elif relative_date+relativedelta(days=i) in list(dataset['Date']):
                    relative_date = relative_date+relativedelta(days=i)
                    return dataset[dataset['Date'].isin([relative_date])].index[0]
                else:
                    continue
            return -1




    def fill_fr(self,name):
        # Read the M3T1PI file to get the security values from the Portfolio selected sheet
        stocks_preds = []
        pct_change_preds = []
        ongoing_preds = []
        path = 'M3T1PI.xlsx'
        files = pd.read_excel(path, sheet_name=None, encoding='unicode_escape', header=None)
        dataset = files['Portfolio Selected']
        sec_values = pd.Series(dataset[~dataset.iloc[:, 1].isnull()].iloc[:, 1], name='security')
        # ongoing_values = pd.Series(dataset[~dataset.iloc[:,5].isnull()].iloc[:,5].unique(), name='ongoing')
        sec_values = sec_values[sec_values != 'Security'].reset_index(drop=True)
        # ongoing_values = ongoing_values[ongoing_values != 'Ongoing'].reset_index(drop=True)

        security_status = sec_values.tolist()

        dataset = files['Portfolio Ongoing']
        pi_po = files['Portfolio Ongoing'] #for later used
        ongoing_values = pd.Series(dataset[~dataset.iloc[:, 1].isnull()].iloc[:, 1], name='ongoing').reset_index(
            drop=True)
        ongoing_values = ongoing_values[3:].tolist()

        net_liq_val = int(float(dataset.iloc[2, 4]))
        Req_equity_as_of_today = dataset.iloc[0, 8]
        Total_Vol_Allowed = dataset.iloc[1, 8]
        Bet_Size_Decided = dataset.iloc[2, 8]
        Total_Heat_Allowed = dataset.iloc[3, 8]
        Max_Margin_per_position = dataset.iloc[4, 8]
        Max_Vol_All_Per_position = dataset.iloc[8, 8] * 100 #i9
        Max_Allowed_Core_Equty = dataset.iloc[6, 8]
        Allwnc_for_Est_pstns = dataset.iloc[7, 8]

        bet_size = files['Portfolio Selected'].iloc[6, 7]
        ongoing_sheet = dataset.copy()

        n_shares = pd.Series(dataset[~dataset.iloc[:, 2].isnull()].iloc[:, 2], name='n_shares').reset_index(drop=True)
        n_shares = n_shares[1:].tolist()


        symbol_lookup = {}
        for index, symbol in enumerate(ongoing_values):
            symbol_lookup[symbol] = n_shares[index]

        # Read the input file
        path = 'M3T1DI.xlsx'
        files = pd.read_excel(path, sheet_name=None, encoding='unicode_escape')

        for key, value in files.items():

            dataset = value.copy()
            dataset.columns = ['Date', 'Close']
            dataset['Date'] = pd.to_datetime(dataset['Date'])
            dataset = dataset.sort_values(by=['Date'], ascending=True).reset_index(drop=True)

            dataset = dataset.set_index('Date')
            dataset.index = pd.to_datetime(dataset.index)

            H3 = dataset.rolling('90 D').max()
            H1 = dataset.rolling('30 D').max()
            L3 = dataset.rolling('90 D').min()
            L1 = dataset.rolling('30 D').min()

            dataset['H3'] = H3
            dataset['H1'] = H1
            dataset['L3'] = L3
            dataset['L1'] = L1

            shift_by_one = dataset[['H3', 'H1', 'L3', 'L1']].shift()
            dataset.drop(['H3', 'H1', 'L3', 'L1'], axis=1, inplace=True)
            dataset = pd.concat([dataset, shift_by_one], axis=1)
            dataset = dataset.reset_index()

            buy = []
            sell = []
            n_shares = []
            p_loss_gain = []
            val_position = []
            entry_exit_loss = []
            account_eq = []
            buy_flag = 0
            sell_flag = 0
            en_ex_gain = 0
            threshold = 90
            acc_eq = 10000

            for idx in range(dataset.shape[0]):

                buy_status = (dataset.loc[idx,'Close'] > dataset.loc[idx,'H3'])
                sell_status = (dataset.loc[idx,'Close'] < dataset.loc[idx,'L1'])

                sell_status_ = (dataset.loc[idx,'Close'] < dataset.loc[idx,'L3'])
                buy_status_ = (dataset.loc[idx,'Close'] > dataset.loc[idx,'H1'])

                if idx < threshold:
                    buy.append('----')
                    sell.append('----')
                    n_shares.append('----')
                    val_position.append('----')
                    entry_exit_loss.append('----')
                    p_loss_gain.append('----')
                    account_eq.append(10000)
                    en_ex_gain = 0
                    continue

                ## first condition
                if (buy_flag == 0) & (buy_status == True) & (sell_flag == 0):
                    buy.append('BUY')
                    sell.append('----')
                    n_shares.append((np.floor(100/(dataset.loc[idx,'Close']-dataset.loc[idx,'L1']))))
                    val_position.append((n_shares[idx]*dataset.loc[idx,'Close']))
                    entry_exit_loss.append('----')
                    p_loss_gain.append('----')
                    account_eq.append(account_eq[idx-1])
                    buy_flag = 1
                elif (buy_flag == 1) & (sell_status == True) & (sell_flag == 0):
                    buy.append('----')
                    sell.append('SELL')
                    n_shares.append(n_shares[idx-1])
                    val_position.append((n_shares[idx]*dataset.loc[idx,'Close']))
                    p_loss_gain.append(val_position[idx]-val_position[idx-1])
                    en_ex_gain += p_loss_gain[idx]
                    entry_exit_loss.append(en_ex_gain)
                    account_eq.append(account_eq[idx-1]+p_loss_gain[idx])
                    buy_flag = 0
                    en_ex_gain = 0
                elif (buy_flag == 1) & (sell_flag == 0):
                    buy.append('----')
                    sell.append('----')
                    n_shares.append(n_shares[idx-1])
                    val_position.append((n_shares[idx]*dataset.loc[idx,'Close']))
                    entry_exit_loss.append('----')
                    p_loss_gain.append(val_position[idx]-val_position[idx-1])
                    account_eq.append(account_eq[idx-1]+p_loss_gain[idx])
                    en_ex_gain += p_loss_gain[idx]

                ## Second condition
                elif (buy_flag == 0) & (sell_status_ == True) & (sell_flag == 0):
                    buy.append('SELL')
                    sell.append('----')
                    n_shares.append((np.ceil(100/(dataset.loc[idx,'Close']-dataset.loc[idx,'H1']))))
                    val_position.append((n_shares[idx]*dataset.loc[idx,'Close']))
                    entry_exit_loss.append('----')
                    p_loss_gain.append('----')
                    account_eq.append(account_eq[idx-1])
                    sell_flag = 1
                elif (buy_flag == 0) & (buy_status_ == True) & (sell_flag == 1):
                    buy.append('----')
                    sell.append('BUY')
                    n_shares.append(n_shares[idx-1])
                    val_position.append((n_shares[idx]*dataset.loc[idx,'Close']))
                    p_loss_gain.append(val_position[idx]-val_position[idx-1])
                    en_ex_gain += p_loss_gain[idx]
                    entry_exit_loss.append(en_ex_gain)
                    account_eq.append(account_eq[idx-1]+p_loss_gain[idx])
                    sell_flag = 0
                    en_ex_gain = 0
                elif (buy_flag == 0) & (sell_flag == 1):
                    buy.append('----')
                    sell.append('----')
                    n_shares.append(n_shares[idx-1])
                    val_position.append((n_shares[idx]*dataset.loc[idx,'Close']))
                    entry_exit_loss.append('----')
                    p_loss_gain.append(val_position[idx]-val_position[idx-1])
                    account_eq.append(account_eq[idx-1]+p_loss_gain[idx])
                    en_ex_gain += p_loss_gain[idx]
                else:
                    buy.append('----')
                    sell.append('----')
                    n_shares.append('----')
                    val_position.append('----')
                    entry_exit_loss.append('----')
                    p_loss_gain.append('----')
                    account_eq.append(account_eq[idx-1])
                    en_ex_gain = 0

            dataset['Enter into position'] = buy
            dataset['Close Position'] = sell
            dataset['How many shares'] = n_shares
            dataset['Value of the position'] = val_position
            dataset['Profit or Loss since previous day'] = p_loss_gain
            dataset['Profit or loss at the exit'] = entry_exit_loss
            dataset['Account Equity'] = account_eq

            KR = dataset['Date']
            current = dataset['Date'][-1:].index[0]
            #mnths_3 = current - 90 - 90
            #mnths_9 = mnths_3 - 180 - 270
            #mnths_12 = mnths_9 - 360 - 630
            difference_dates = max(KR).date() - min(KR).date()
            if difference_dates.days > 65:
                mnths_3 = self.extract_index(dataset['Date'][-1:].iloc[0], 3, dataset)
            else:
                mnths_3 = -1
            if difference_dates.days > 250:
                mnths_9 = self.extract_index(dataset['Date'].iloc[mnths_3], 6, dataset)
            else:
                mnths_9 = -1
            if difference_dates.days > 635:
                mnths_12 = self.extract_index(dataset['Date'].iloc[mnths_9], 12, dataset)
            else:
                mnths_12 = -1


            pct_change = []
            cur_val = account_eq[-1:][0]
            days_range = [[mnths_3, current], [mnths_9, mnths_3], [mnths_12, mnths_9]]

            for idx in range(3):
                if days_range[idx][0] != -1:
                    max_month = max(account_eq[days_range[idx][0]:days_range[idx][1]])
                    min_month = min(account_eq[days_range[idx][0]:days_range[idx][1]])

                    pct_change_x = ((max_month-cur_val)/max_month)*100
                    pct_change_y = ((cur_val-min_month)/min_month)*100

                    pct_change.append(pct_change_x-pct_change_y)
                else:
                    pct_change.append("NA")

            pct_change = pd.DataFrame(pct_change).T
            pct_change.columns = ['3_months', '6_months', '12_months']


            ## Condition to check whether trade is ongoing on not
            if key in security_status:
                if key in ongoing_values:
                    pct_change['Ongoing'] = 'YES'
                    pct_change['Buy_or_Sell'] = '----'
                    pct_change['H3'] = '----'
                    pct_change['L3'] = '----'
                    pct_change['H1'] = '----'
                    pct_change['L1'] = '----'
                    pct_change['C'] = '----'
                    pct_change['How much from Engine'] = '----'

                    exit_price = dataset['L1'][-1:].values[0]
                    close_price = dataset['Close'][-1:].values[0]
                    heat = 0

                    if symbol_lookup[key] < 0:
                        exit_price = dataset['H1'][-1:].values[0]
                        heat = abs(symbol_lookup[key] * (exit_price-close_price))
                    else:
                        heat = abs(symbol_lookup[key] * (close_price-exit_price))
                    vol = '{:.4f}%'.format((heat / net_liq_val) * 100)
                    ongoing_preds.append([key, symbol_lookup[key], exit_price, close_price, abs(heat), vol])

                else:
                    pct_change['Ongoing'] = 'NO'
                    if dataset['Enter into position'][-1:].values[0] != '----':
                        pct_change['Buy_or_Sell'] = dataset['Enter into position'][-1:].values[0]
                        pct_change['H3'] = dataset['H3'][-1:].values[0]
                        pct_change['L3'] = dataset['L3'][-1:].values[0]
                        pct_change['H1'] = dataset['H1'][-1:].values[0]
                        pct_change['L1'] = dataset['L1'][-1:].values[0]
                        pct_change['C'] = dataset['Close'][-1:].values[0]
                        if dataset['Enter into position'][-1:].values[0] == 'BUY':
                            pct_change['How much from Engine'] = bet_size / (dataset['Close'][-1:].values[0]-dataset['L1'][-1:].values[0])
                        else:
                            pct_change['How much from Engine'] = bet_size / (dataset['H1'][-1:].values[0]-dataset['Close'][-1:].values[0])
                    else:
                        pct_change['Buy_or_Sell'] = '----'
                        pct_change['H3'] = '----'
                        pct_change['L3'] = '----'
                        pct_change['H1'] = '----'
                        pct_change['L1'] = '----'
                        pct_change['C'] = '----'
                        pct_change['How much from Engine'] = '----'
            else:
                pct_change['Ongoing'] = '----'
                pct_change['Buy_or_Sell'] = '----'
                pct_change['H3'] = '----'
                pct_change['L3'] = '----'
                pct_change['H1'] = '----'
                pct_change['L1'] = '----'
                pct_change['C'] = '----'
                pct_change['How much from Engine'] = '----'

            pct_change = pct_change.rename(index={0: key})

            stocks_preds.append(dataset)
            pct_change_preds.append(pct_change)


        ongoing_preds = pd.DataFrame(ongoing_preds)
        ongoing_preds.columns = ['Security', 'How many', 'Exit price', 'Closing price', 'Heat', 'Vol']


        total_heat = np.sum(ongoing_preds['Heat'])
        Totl_Vol_as_of_today = total_heat / net_liq_val


        ongoing_preds['Buy_or_Sell'] = '----'
        # for idx in range(ongoing_preds.shape[0]):
        #     # print(symbol_lookup[ongoing_preds['Security'][idx]])
        #     if (symbol_lookup[ongoing_preds['Security'][idx]] > 0) & ((ongoing_preds['Vol'][idx] > Max_Vol_All_Per_position) & (Totl_Vol_as_of_today > Total_Vol_Allowed)):
        #         ongoing_preds.iloc[idx, 6] = "SELL"
        #     elif (symbol_lookup[ongoing_preds['Security'][idx]] < 0) & ((ongoing_preds['Vol'][idx] < Max_Vol_All_Per_position) & (Totl_Vol_as_of_today > Total_Vol_Allowed)):
        #         ongoing_preds.iloc[idx, 6] = "BUY"

            # IF(A, "SELL", B)
            # A = AND(C14 > 0, G14 >$I$9,  $N$5 > 0  )
            # B = IF(AND(C14 < 0, G14 >$I$9,$N$5 > 0), "BUY", "")


        ongoing_sheet.iloc[6,4] = np.sum(ongoing_preds['Heat'])
        ongoing_sheet.iloc[7,4] = Totl_Vol_as_of_today
        row = 13
        col = 1


        for idx_1 in range(ongoing_preds.shape[0]):
            for idx_2 in range(ongoing_preds.shape[1]):
                ongoing_sheet.iloc[row, col] = ongoing_preds.iloc[idx_1, idx_2]
                col = col+1
            col = 1
            row = row+1

        p_select = pd.DataFrame()
        for idx in range(len(pct_change_preds)):
            p_select = pd.concat([p_select, pct_change_preds[idx].reset_index()], axis=0).reset_index(drop=True)
        p_select.columns = ['symbol', '3_months', '6_months', '12_months', 'Ongoing', 'Buy_or_Sell', 'H3', 'L3', 'H1', 'L1', 'C', 'How much from Engine']


        p_selected = p_select[p_select.symbol.isin(sec_values)].reset_index(drop=True)
        p_selection = p_select[['symbol', '3_months', '6_months', '12_months']]

        path = name
        m3 = float(pi_po.iloc[2, 11])
        m6 = float(pi_po.iloc[2, 12])
        m12 = float(pi_po.iloc[2, 13])
        x = [float(pi_po.iloc[5, 11]),
             float(pi_po.iloc[5, 12]),
             float(pi_po.iloc[5, 13]),
             float(pi_po.iloc[5, 14]),
             float(pi_po.iloc[5, 15]),
             float(pi_po.iloc[5, 16]),
             float(pi_po.iloc[6, 11]),
             float(pi_po.iloc[6, 12]),
             float(pi_po.iloc[6, 13]),
             float(pi_po.iloc[6, 14]),
             float(pi_po.iloc[6, 15]),
             float(pi_po.iloc[6, 16]),
             float(pi_po.iloc[7, 11]),
             float(pi_po.iloc[7, 12]),
             float(pi_po.iloc[7, 13]),
             float(pi_po.iloc[7, 14]),
             float(pi_po.iloc[7, 15]),
             float(pi_po.iloc[7, 16])]
        value1 = float(pi_po.iloc[2, 16])
        value2 = float(pi_po.iloc[3, 16])
        p_selected = p_selected.replace(r'NA', '0')
        first = pd.DataFrame(
            columns=['symbol', '3_months', '6_months', '12_months', 'Ongoing', 'Buy_or_Sell', 'H3', 'L3', 'H1', 'L1',
                     'C', 'How much from Engine'])
        second = pd.DataFrame(
            columns=['symbol', '3_months', '6_months', '12_months', 'Ongoing', 'Buy_or_Sell', 'H3', 'L3', 'H1', 'L1',
                     'C', 'How much from Engine'])
        third = pd.DataFrame(
            columns=['symbol', '3_months', '6_months', '12_months', 'Ongoing', 'Buy_or_Sell', 'H3', 'L3', 'H1', 'L1',
                     'C', 'How much from Engine'])
        fourth = pd.DataFrame(
            columns=['symbol', '3_months', '6_months', '12_months', 'Ongoing', 'Buy_or_Sell', 'H3', 'L3', 'H1', 'L1',
                     'C', 'How much from Engine'])
        for i in range(0, len(x)):
            if math.isnan(x[i]):
                if i % 2 == 0:
                    x[i] = -math.inf
                else:
                    x[i] = math.inf
        for index, row in p_selected.iterrows():
            if float(row["3_months"]) >= m3 and float(row["6_months"]) >= m6 and float(row["12_months"]) >= m12:
                if float(row["3_months"]) >= x[0] and float(row["6_months"]) >= x[2] and float(row["12_months"]) >= x[4]:
                    if float(row["3_months"]) <= x[1] and float(row["6_months"]) <= x[3] and float(row["12_months"]) <= x[5]:
                        first = first.append(row, ignore_index=True)
                        continue
                if float(row["3_months"]) >= x[6] and float(row["6_months"]) >= x[8] and float(row["12_months"]) >= x[10]:
                    if float(row["3_months"]) <= x[7] and float(row["6_months"]) <= x[9] and float(row["12_months"]) <= x[11]:
                        second = second.append(row, ignore_index=True)
                        continue
                if float(row["3_months"]) >= x[12] and float(row["6_months"]) >= x[14] and float(row["12_months"]) >= x[16]:
                    if float(row["3_months"]) <= x[13] and float(row["6_months"]) <= x[15] and float(row["12_months"]) <= x[17]:
                        third = third.append(row, ignore_index=True)
                        continue
                fourth = fourth.append(row, ignore_index=True)



        first.sort_values(["12_months", "6_months", "3_months"], ascending=(False, False, False), inplace=True)
        second.sort_values(["3_months", "6_months", "12_months"], ascending=(False, False, False), inplace=True)
        third.sort_values(["6_months", "3_months", "12_months"], ascending=(False, False, False), inplace=True)
        fourth.sort_values(["3_months", "6_months", "12_months"], ascending=(False, False, False), inplace=True)
        p_selected = pd.concat([first, second, third, fourth])


        n_shares = pd.Series(ongoing_sheet[~ongoing_sheet.iloc[:, 2].isnull()].iloc[:, 2], name='n_shares').reset_index(drop=True)
        total_positions = len(n_shares[1:].tolist())
        ans_t_13 = 0


        #clear Fr file
        ongoing_sheet.iloc[1:9, 10:24] = float("NaN")
        ongoing_sheet.iloc[0, 10] = "How many positions in the portfolio"
        ongoing_sheet.iloc[0, 13] = total_positions
        ongoing_sheet.iloc[1, 10] = "What is the core equIty shoUld be according to max vol allowed for the whole portfolio per position(Based on NLV)"
        ongoing_sheet.iloc[1, 13] = float(ongoing_sheet.iloc[2,4])*(1-(float(ongoing_sheet.iloc[0,13])*(float(ongoing_sheet.iloc[1,8])+float(ongoing_sheet.iloc[3,8]))/(float(ongoing_sheet.iloc[3,8])/float(ongoing_sheet.iloc[2,8]))))
        ongoing_sheet.iloc[2, 10] = "What is the allowed core equity should be according to allowed heat, Req as of today and no of instruments in the portforlio(Based on Req as of today)"
        ongoing_sheet.iloc[2, 13] = max(float(ongoing_sheet.iloc[6,8]),(float(ongoing_sheet.iloc[6,8])+(float(ongoing_sheet.iloc[0,8])-float(ongoing_sheet.iloc[6,8]))*(1-(float(ongoing_sheet.iloc[0,13])/(float(ongoing_sheet.iloc[3,8])/float(ongoing_sheet.iloc[2,8]))))))
        ongoing_sheet.iloc[4, 10] = "How much of heat should be reduced to control Vol"
        ongoing_sheet.iloc[5, 10] = "Vol after account for min vol"


        ongoing_sheet.iloc[6, 10] = "New position bet size"
        ongoing_sheet.iloc[6, 12] = int()
        ongoing_sheet.iloc[7, 10] = "Pyramid amount"
        ongoing_sheet.iloc[7, 12] = int()
        sum_l6 = 0
        sum_m6 = 0
        if float(ongoing_sheet.iloc[2, 4]) > float(ongoing_sheet.iloc[0, 8]):
            ans_t_13 = max(0, (float(ongoing_sheet.iloc[1, 13]) - (
                        float(ongoing_sheet.iloc[2, 4]) - float(ongoing_sheet.iloc[6, 4]))))
        else:
            ans_t_13 = max(0, min((float(ongoing_sheet.iloc[2, 13]) - (
                        float(ongoing_sheet.iloc[2, 4]) - float(ongoing_sheet.iloc[6, 4]))), (
                                              float(ongoing_sheet.iloc[6, 4]) - (
                                                  float(ongoing_sheet.iloc[2, 4]) * float(ongoing_sheet.iloc[2, 8])))))
        for index_i in range(0, total_positions):
            ongoing_sheet.iloc[13+index_i, 7] = '----'
            g_col_value = float(ongoing_sheet.iloc[13+index_i, 6].replace("%",""))
            if (ongoing_sheet.iloc[13+index_i, 2] > 0) & ((g_col_value> Max_Vol_All_Per_position) & (ans_t_13 > 0)):
                ongoing_sheet.iloc[13+index_i, 7] = "SELL"
            elif (ongoing_sheet.iloc[13+index_i, 2] < 0) & ((g_col_value > Max_Vol_All_Per_position) & (ans_t_13 > 0)):
                ongoing_sheet.iloc[13+index_i, 7] = "BUY"

        for i_total_position in range(0,total_positions):
            g_col_value = float(ongoing_sheet.iloc[13 + i_total_position, 6].replace("%", ""))
            if ongoing_sheet.iloc[13 + i_total_position, 7] == "SELL":
                sum_l6 = sum_l6 + g_col_value
            if ongoing_sheet.iloc[13 + i_total_position, 7] == "BUY":
                sum_m6 = sum_m6 + g_col_value

        ongoing_sheet.iloc[5, 11] = '{:.2f}%'.format(sum_l6)
        ongoing_sheet.iloc[5, 12] = '{:.2f}%'.format(sum_m6)
        ongoing_sheet.iloc[5, 13] = '{:.2f}%'.format(sum_l6+ sum_m6)
        ongoing_sheet.iloc[4, 13] = ans_t_13
        ongoing_sheet.iloc[3, 7] = "Total Heat allowed for the portfolio"

        for index_i in range(0, total_positions):
            ongoing_sheet.iloc[13 + index_i, 8] = int()

            g_col_value = float(ongoing_sheet.iloc[13 + index_i, 6].replace("%", ""))
            if ongoing_sheet.iloc[13 + i_total_position, 7] in ["SELL","BUY"]:
                if (g_col_value > Max_Vol_All_Per_position) & (ans_t_13 > 0):
                    if g_col_value > Max_Vol_All_Per_position:
                        ongoing_sheet.iloc[13+index_i, 8] = abs(int(((g_col_value / float(ongoing_sheet.iloc[5, 13].replace("%","")))*ans_t_13) * (ongoing_sheet.iloc[13 + index_i, 2]/ongoing_sheet.iloc[13 + index_i, 5])))
                elif (g_col_value > ongoing_sheet.iloc[5, 8]) & (ans_t_13 == 0):
                    if g_col_value > (ongoing_sheet.iloc[5, 8]+ongoing_sheet.iloc[8, 8]):
                        ongoing_sheet.iloc[13+index_i, 8] = abs((ongoing_sheet.iloc[13+index_i, 2] * ongoing_sheet.iloc[8,8])/g_col_value)
        e14 = total_heat
        e9 = net_liq_val - e14
        j14 = 0
        if e9 < ongoing_sheet.iloc[2, 8]:
            j14 = 0
        else:
            j14 = max(0,e9-ongoing_sheet.iloc[2, 8])

        j15 = max((net_liq_val - ongoing_sheet.iloc[0, 8]),0)
        j17 = 0
        if net_liq_val > ongoing_sheet.iloc[0, 8]:
            j17 = max(( e9 - ongoing_sheet.iloc[1, 13]),0)
        else:
            j17 = 0
        j18 = 0
        if (net_liq_val - ongoing_sheet.iloc[0, 8]) > 0:
            j18 = min(j17,j15)
        else:
            j18 = 0
        j16 = 0
        if  ongoing_sheet.iloc[4, 13]> 0:
            j16 = 0
        else:
            j16 = max(min(j14,(j17-j18)),0)
        j19 = j16+ j18
        l13 = 0
        n14 = e9 - ongoing_sheet.iloc[6, 8]
        n15 = ongoing_sheet.iloc[0, 8] - ongoing_sheet.iloc[6, 8]
        n16 = ongoing_sheet.iloc[3, 8] - ongoing_sheet.iloc[2, 8]
        n17 = n16 - ongoing_sheet.iloc[0, 13]
        n19 = (n17 / n16) * n15
        n20 = n19 / n17
        n23 = ((n17/n16)*(n14-j19 * ongoing_sheet.iloc[9, 8]))/n17+ongoing_sheet.iloc[4, 13]/n17
        n24 = 0
        n22 = net_liq_val * ongoing_sheet.iloc[2, 8]
        if net_liq_val < ongoing_sheet.iloc[0, 8]:
            n24 = n22
        else:
            n24 = max(n23,n20)
        # ((N17 / N16) * (N14 - J19 * L13)) / N17 + J9 / N17
        # ((4 / 5) * (325 - J19 * L13)) / N17 + J9 / N17
        ongoing_sheet.iloc[6, 12] = n24
        ongoing_sheet.iloc[7, 12] = j19



        with pd.ExcelWriter(path) as writer:
            ongoing_sheet.to_excel(writer, sheet_name='Portfolio Ongoing', index=False, header=None)
            p_selected.to_excel(writer, sheet_name='Portfolio Selected', index=False, startrow=8, startcol=1)
            p_selection.to_excel(writer, sheet_name='Portfolio Selection', index=False, startrow=8, startcol=1)
            for idx in range(len(pct_change_preds)):
                stocks_preds[idx].to_excel(writer, sheet_name=p_select['symbol'][idx], index=False)
                
def fr_file():
    if os.path.exists('M3T1fr.xlsx'):
        os.remove('M3T1fr.xlsx')
        print("Old Fr file deleted")
    obj = create_fr()
    obj.fill_fr('M3T1fr.xlsx')
    print("FR file Created")

def new_fr_file():
    if os.path.exists('NEW_fr.xlsx'):
        os.remove('NEW_fr.xlsx')
    obj = create_fr()
    obj.fill_fr('NEW_fr.xlsx')
    print("New FR file Created")


# In[7]:

def od_files_download():
    profile_data.profile_info()
    placed_order_history()

def data_download_complete_pi():
#     if os.path.exists("m3t2OD1.xlsx"):
#         os.remove("m3t2OD1.xlsx")
#     if os.path.exists("m3t2OD2.xlsx"):
#         os.remove("m3t2OD2.xlsx")
#     if os.path.exists("m3t2OD3.xlsx"):
#         os.remove("m3t2OD3.xlsx")
#     profile_data.profile_info()
#     placed_order_history()
    clean_PI_file()
    create_pi_file("M3T1PI.xlsx")

def delete_extra():
    if os.path.exists('Complete_Output.xlsx'):
        os.remove('Complete_Output.xlsx')
    print("File Complete_Output Deleted ")
    if os.path.exists('M3T1DI.xlsx'):
        os.remove('M3T1DI.xlsx')
    print("File M3T1DI Deleted ")
    if os.path.exists("m3t2OD1.xlsx"):
        os.remove("m3t2OD1.xlsx")
    print("File m3t2OD1 Deleted ")
    if os.path.exists("m3t2OD2.xlsx"):
        os.remove("m3t2OD2.xlsx")
    print("File m3t2OD2 Deleted ")
    if os.path.exists("m3t2OD3.xlsx"):
        os.remove("m3t2OD3.xlsx")
    print("File m3t2OD3 Deleted ")
    if os.path.exists('M3T1fr.xlsx'):
        os.remove('M3T1fr.xlsx')
    print("File M3T1fr Deleted ")
    if os.path.exists('new_place_order.xlsx'):
        os.remove('new_place_order.xlsx')
    print("File new_place_order Deleted ")
    if os.path.exists('stoploss_place_order.xlsx'):
        os.remove('stoploss_place_order.xlsx')
    print("File stoploss_place_order Deleted ")


# In[ ]:

##New place order file code commented as without stoploss this will be updated from fr selected
# class new_place_order_file():
#     def n_p_o_f(self):
#         path= 'Place_order_inputs.xlsx'
#         po = pd.read_excel(path, encoding = 'unicode_escape', header=None)
#         path= 'Historical_data_inputs.xlsx'
#         hd = pd.read_excel(path, encoding = 'unicode_escape')
#         path= 'M3T1fr.xlsx'
#         fr = pd.read_excel(path, sheet_name=None, encoding = 'unicode_escape', header=None)
#
#         Port_folio = fr['Portfolio Ongoing']
#         Port_folio_S = fr['Portfolio Selected']
#         #from fr file portfolio ongoing
#         i=13
#         sym=[]
#         # ext=[]
#         ot=[]
#         # act=[]
#         tq=[]
#         lt=[]
#         try:
#             while(str(Port_folio.iloc[i,1]) != 'nan'):
#                 sym.append(Port_folio.iloc[i,1])
#                 # ext.append(Port_folio.iloc[i,3])
#                 # act.append(Port_folio.iloc[i,7])
#                 tq.append(Port_folio.iloc[i,8])
#                 if Port_folio.iloc[i,7] == "BUY":
#                     lt.append(Port_folio.iloc[i,3]*0.99)
#                 else:
#                     lt.append(Port_folio.iloc[i, 3] * 1.01)
#                 ot.append("STPLMT")
#                 i=i+1
#         except:
#             print("Fr file Read Successful")
#
#         #from hist data
#         st=[]
#         exc=[]
#         cur=[]
#         pexc=[]
#         for i in range(0,len(sym)): #fr file and Hist data file
#             for j in range(0,len(hd.symbol)):
#                 if(sym[i] ==  hd.symbol[j]):
#                     st.insert(i,hd.secType[j])
#                     exc.insert(i,hd.exchange[j])
#                     cur.insert(i,hd.currency[j])
#                     pexc.insert(i,hd.primaryExchange[j])
#                     continue
#
#         #saving data
#         new_po=pd.DataFrame()
#         new_po["symbol"]=sym
#         new_po["secType"]=st
#         new_po["exchange"]=exc
#         new_po["currency"]=cur
#         new_po["primaryExchange"]=pexc
#         new_po["action"]=act
#         new_po["totalQuantity"]=tq
#         new_po["orderType"]=ot
#         new_po["lmtPrice"]=lt
#         new_po["stpPrice"]=ext
#
#         # new_po.to_excel()
#         init_output_file("new_place_order.xlsx",1)
#         output_writer("new_place_order.xlsx",new_po,"New_place_order",0)
#         init_output_file("new_place_order.xlsx",2)
#
# def new_place_order():
#     obj=new_place_order_file()
#     obj.n_p_o_f()
#     print("New Place order file created")


class new_place_order_file():
    def n_p_o_f(self):
        path= 'Place_order_inputs.xlsx'
        po = pd.read_excel(path, encoding = 'unicode_escape', header=None)
        path= 'Historical_data_inputs.xlsx'
        hd = pd.read_excel(path, encoding = 'unicode_escape')
        path= 'M3T1fr.xlsx'
        fr = pd.read_excel(path, sheet_name=None, encoding = 'unicode_escape', header=None)

        Port_folio = fr['Portfolio Ongoing']
        Port_folio_S = fr['Portfolio Selected']
        #from fr file portfolio ongoing
        # Port_folio_S = Port_folio_S.replace(r'----', '0')
        i=9
        sym=[]
        ext=[]
        ot=[]
        act=[]
        tq=[]
        lt=[]
        try:
            while(str(Port_folio_S.iloc[i,1]) != 'nan'):
                sym.append(Port_folio_S.iloc[i,1])
                ext.append(Port_folio_S.iloc[i,11])
                act.append(Port_folio_S.iloc[i,6])
                if Port_folio_S.iloc[i,12] == "----":
                    tq.append(str(0))
                else:
                    tq.append(str(int(Port_folio_S.iloc[i, 12])))
                temp_type = check_int_float(Port_folio_S.iloc[i,11])
                if temp_type != "False":
                    if Port_folio_S.iloc[i,6] == "BUY":
                        lt.append(temp_type*0.99)
                    else:
                        lt.append(temp_type * 1.01)
                else:
                    lt.append("----")
                ot.append("STPLMT")
                i=i+1
        except Exception as e:

            print(e)
            print("Fr file Read Successful")

        #from hist data
        st_new =[]
        exc_new =[]
        cur_new =[]
        pexc_new =[]
        sym_new = []
        ext_new = []
        ot_new = []
        act_new = []
        tq_new = []
        lt_new = []
        index_i = 0
        for i in range(0,len(sym)): #fr file and Hist data file
            for j in range(0,len(hd.symbol)):
                if(sym[i] ==  hd.symbol[j]):
                    sym_new.insert(index_i,sym[i])
                    ext_new.insert(index_i,ext[i])
                    ot_new.insert(index_i,ot[i])
                    act_new.insert(index_i,act[i])
                    tq_new.insert(index_i,tq[i])
                    lt_new.insert(index_i,lt[i])
                    st_new.insert(index_i,hd.secType[j])
                    exc_new.insert(index_i,hd.exchange[j])
                    cur_new.insert(index_i,hd.currency[j])
                    pexc_new.insert(index_i,hd.primaryExchange[j])
                    index_i = index_i + 1
                    continue

        #saving data
        new_po=pd.DataFrame()
        new_po["symbol"]=sym_new
        new_po["secType"]=st_new
        new_po["exchange"]=exc_new
        new_po["currency"]=cur_new
        new_po["primaryExchange"]=pexc_new
        new_po["action"]=act_new
        new_po["totalQuantity"]=tq_new
        new_po["orderType"]=ot_new
        new_po["lmtPrice"]=lt_new
        new_po["stpPrice"]=ext_new

        # new_po.to_excel()
        init_output_file("new_place_order.xlsx",1)
        output_writer("new_place_order.xlsx",new_po,"New_place_order",0)
        init_output_file("new_place_order.xlsx",2)

def new_place_order():
    obj=new_place_order_file()
    obj.n_p_o_f()
    print("New Place order file created")

class stoploss_place_order_file():
    def s_p_o_f(self):
        path= 'Historical_data_inputs.xlsx'
        hd = pd.read_excel(path, encoding = 'unicode_escape')
        path= 'M3T1fr.xlsx'
        fr = pd.read_excel(path, sheet_name=None, encoding = 'unicode_escape', header=None)

        Port_folio = fr['Portfolio Ongoing']
        Port_folio_S = fr['Portfolio Selected']
        #from fr file portfolio ongoing
        i=13
        sym=[]
        act = []
        tq = []
        ot = []
        ext=[]
        lt=[]

        try:
            while(str(Port_folio.iloc[i,1]) != 'nan'):
                sym.append(Port_folio.iloc[i,1])
                tq.append(Port_folio.iloc[i, 2])
                if Port_folio.iloc[i, 2] > 0:
                    act.append("SELL")
                else:
                    act.append("BUY")
                ot.append("STP")
                ext.append(Port_folio.iloc[i,3])
                lt.append("---")
                i=i+1
        except:
            print("Fr file Read Successful")

        #from hist data
        st=[]
        exc=[]
        cur=[]
        pexc=[]
        for i in range(0,len(sym)): #fr file and Hist data file
            for j in range(0,len(hd.symbol)):
                if(sym[i] ==  hd.symbol[j]):
                    st.insert(i,hd.secType[j])
                    exc.insert(i,hd.exchange[j])
                    cur.insert(i,hd.currency[j])
                    pexc.insert(i,hd.primaryExchange[j])
                    continue

        #saving data
        stoploss_po=pd.DataFrame()
        stoploss_po["symbol"]=sym
        stoploss_po["secType"]=st
        stoploss_po["exchange"]=exc
        stoploss_po["currency"]=cur
        stoploss_po["primaryExchange"]=pexc
        stoploss_po["action"]=act
        stoploss_po["totalQuantity"]=tq
        stoploss_po["orderType"]=ot
        stoploss_po["lmtPrice"]=lt
        stoploss_po["stpPrice"]=ext

        # stoploss_po.to_excel()
        init_output_file("stoploss_place_order.xlsx",1)
        output_writer("stoploss_place_order.xlsx",stoploss_po,"stoploss_place_order",0)
        init_output_file("stoploss_place_order.xlsx",2)

def stoploss_place_order():
    obj=stoploss_place_order_file()
    obj.s_p_o_f()
    print("Stoploss Place order file created")


def re_update_ui():
    map = {}
    path = 'M3T1PI.xlsx'
    pi = pd.read_excel(path, sheet_name=None, encoding='unicode_escape', header=None)
    path = 'M3T1fr.xlsx'
    fr = pd.read_excel(path, sheet_name=None, encoding='unicode_escape', header=None)
    on_pi = pi['Portfolio Ongoing']
    sel_pi = pi['Portfolio Selected']
    on_fr = fr['Portfolio Ongoing']
    on_pi.iloc[6, 4] = on_fr.iloc[6, 4]  # total heat
    output_writer_headers('M3T1PI.xlsx', on_pi, 'Portfolio Ongoing', 1)
    print("Total Heat updated in PI file")
    map["NLV                        "] = round(float(on_pi.iloc[2, 4]), 2)
    map["Req as of today     "] = round(float(on_pi.iloc[0, 8]), 2)
    map["Bet size                   "] = round(float(on_pi.iloc[2, 8]), 2)
    map["Bet amount           "] = round(float(sel_pi.iloc[6, 7]), 2)
    map["Total heat               "] = round(float(on_fr.iloc[6, 4]), 2)
    map["Total Vol allowed  "] = round(float(on_fr.iloc[7, 4]), 2)

    new_box(map)

# In[9]:
root = Tk()
style = ttk.Style()
root.title('Communicator - Interactive Broker')
root.rowconfigure(10, weight=1)
root.columnconfigure(0, weight=1)
root.geometry('100x100')
style = Style() 
style.configure('W.TButton', font =('calibri', 10), foreground = 'red')
MyButton1 = Button(root, text="Download Historical Data", width=25, command=data_download)
MyButton1.place(x=30, y=50)
MyButton1 = Button(root, text="Clean Data", width=20, command=hist_data_formating)
MyButton1.place(x=220, y=50)
MyButton1 = Button(root, text="Download od files", width=52, command=od_files_download)
MyButton1.place(x=30, y=80)
MyButton1 = Button(root, text="Update PI File", width=52, command=data_download_complete_pi)
MyButton1.place(x=30, y=110)
MyButton1 = Button(root, text="Prepare FR File", width=52, command=fr_file)
MyButton1.place(x=30, y=150)
MyButton1 = Button(root, text="Update Place Orders inputs", width=25, command=new_place_order)
MyButton1.place(x=30, y=200)
MyButton1 = Button(root, text="Place Orders", width=20, command=order_placer)
MyButton1.place(x=220, y=200)
MyButton1 = Button(root, text="Update StopLoss Orders inputs", width=27, command=stoploss_place_order)
MyButton1.place(x=30, y=250)
MyButton1 = Button(root, text="Modify StopLoss", width=20, command=stoploss_order_placer)
MyButton1.place(x=220, y=250)
MyButton1 = Button(root, text="Reupdate PI File", width=52, command=re_update_ui)
MyButton1.place(x=30, y=300)
MyButton1 = Button(root, text="Redo FR File", width=52, command=new_fr_file)
MyButton1.place(x=30, y=350)

MyButton1 = Button(root, text="Delete Extra Files", width=30, command=delete_extra)
MyButton1.place(x=100, y=400)
MyButton1 = Button(root, text="Exit Application", width=15,style = 'W.TButton',command=root.destroy)
MyButton1.place(x=250, y=550)
center_window(400, 600)

mainloop()







